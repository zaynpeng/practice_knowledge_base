from __future__ import annotations

import csv
import json
import mimetypes
import os
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
try:
    from bs4 import BeautifulSoup
except ImportError:  # The app can still run; extraction falls back to a basic cleaner.
    BeautifulSoup = None
from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "knowledge.db"
SCHEMA_PATH = BASE_DIR / "schema.sql"
ATTACHMENTS_DIR = BASE_DIR / "attachments"
EXPORTS_DIR = BASE_DIR / "exports"
BACKUPS_DIR = BASE_DIR / "backups"

CONTENT_STATUSES = ["待补充", "待判断", "待实践", "实践中", "等待复盘", "已验证", "不适用", "已沉淀"]
EXPERIMENT_STATUSES = ["待实践", "实践中", "等待复盘", "已验证", "不适用", "已沉淀"]
PROBLEM_STATUSES = ["待处理", "正在探索", "正在实践", "已有方案", "暂不处理", "已解决"]
BOOK_RELATIONSHIPS = ["", "当前需要", "近期可读", "参考工具书", "同类重复", "兴趣阅读", "暂不投入"]
READING_STATUSES = ["未开始", "准备阅读", "阅读中", "暂停", "已完成", "仅查阅", "放弃"]
PRIORITIES = ["", "高", "中", "低", "未判断"]

app = Flask(__name__)
app.secret_key = "local-practice-knowledge-base"


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_dirs() -> None:
    for folder in [
        DATA_DIR,
        ATTACHMENTS_DIR,
        EXPORTS_DIR / "markdown" / "内容收藏",
        EXPORTS_DIR / "markdown" / "读书感悟",
        EXPORTS_DIR / "markdown" / "实践记录",
        EXPORTS_DIR / "markdown" / "问题地图",
        EXPORTS_DIR / "markdown" / "已沉淀方法",
        EXPORTS_DIR / "csv",
        EXPORTS_DIR / "json",
        BACKUPS_DIR,
    ]:
        folder.mkdir(parents=True, exist_ok=True)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error=None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    ensure_dirs()
    with sqlite3.connect(DB_PATH) as db:
        db.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))


def row_to_dict(row: sqlite3.Row | None) -> dict | None:
    return dict(row) if row else None


def query_all(sql: str, params: tuple = ()) -> list[sqlite3.Row]:
    return get_db().execute(sql, params).fetchall()


def query_one(sql: str, params: tuple = ()) -> sqlite3.Row | None:
    return get_db().execute(sql, params).fetchone()


def execute(sql: str, params: tuple = ()) -> int:
    db = get_db()
    cur = db.execute(sql, params)
    db.commit()
    return cur.lastrowid


def form_value(name: str) -> str:
    return request.form.get(name, "").strip()


def content_status(save_reason: str, wants_practice: bool) -> str:
    if not save_reason:
        return "待补充"
    return "待实践" if wants_practice else "待判断"


def unique_path(folder: Path, name: str) -> Path:
    safe_name = secure_filename(name) or "export.md"
    path = folder / safe_name
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    return folder / f"{stem}_{stamp()}{suffix}"


def extract_url(url: str) -> dict:
    if not url:
        return {"status": "仅保存链接", "error": "", "title": "", "raw_text": "", "summary": ""}
    try:
        response = requests.get(url, timeout=8, headers={"User-Agent": "PracticeKnowledgeBase/1.0"})
        response.raise_for_status()
        response.encoding = response.apparent_encoding or response.encoding
        if BeautifulSoup is None:
            text = response.text
            title = ""
            lower = text.lower()
            if "<title" in lower and "</title>" in lower:
                start = lower.find(">", lower.find("<title")) + 1
                end = lower.find("</title>", start)
                title = text[start:end].strip()[:300]
            raw_text = " ".join(text.replace("<", " <").replace(">", "> ").split())
            return {"status": "部分读取成功", "error": "未安装 beautifulsoup4，仅完成基础文本读取。", "title": title, "raw_text": raw_text[:30000], "summary": raw_text[:450]}
        soup = BeautifulSoup(response.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "aside", "form"]):
            tag.decompose()
        title = (soup.title.string.strip() if soup.title and soup.title.string else "")[:300]
        main = soup.find("article") or soup.find("main") or soup.body or soup
        paragraphs = [p.get_text(" ", strip=True) for p in main.find_all(["p", "h1", "h2", "h3", "li"])]
        raw_text = "\n".join([p for p in paragraphs if len(p) > 12])
        if not raw_text:
            return {"status": "部分读取成功", "error": "页面可访问，但未提取到正文。", "title": title, "raw_text": "", "summary": ""}
        summary = raw_text[:450] + ("..." if len(raw_text) > 450 else "")
        return {"status": "读取成功", "error": "", "title": title, "raw_text": raw_text[:50000], "summary": summary}
    except Exception as exc:
        return {"status": "读取失败", "error": str(exc), "title": "", "raw_text": "", "summary": ""}


def save_attachments(entity_type: str, entity_id: int) -> None:
    files = request.files.getlist("attachments")
    for file in files:
        if not file or not file.filename:
            continue
        folder = ATTACHMENTS_DIR / entity_type / str(entity_id)
        folder.mkdir(parents=True, exist_ok=True)
        path = unique_path(folder, file.filename)
        file.save(path)
        execute(
            "INSERT INTO attachments(entity_type, entity_id, file_name, file_path, mime_type, created_at) VALUES(?,?,?,?,?,?)",
            (entity_type, entity_id, file.filename, str(path.relative_to(BASE_DIR)), file.mimetype or mimetypes.guess_type(path.name)[0], now()),
        )


def markdown_filename(title: str, fallback: str = "未命名") -> str:
    safe = "".join(ch for ch in (title or fallback) if ch not in r'\/:*?"<>|').strip()
    return f"{safe[:60] or fallback}_{stamp()}.md"


def write_markdown(subdir: str, title: str, body: str) -> Path:
    folder = EXPORTS_DIR / "markdown" / subdir
    folder.mkdir(parents=True, exist_ok=True)
    path = unique_path(folder, markdown_filename(title))
    path.write_text(body, encoding="utf-8")
    return path


def existing_problem_links(entity: str, entity_id: int) -> list[str]:
    table = {"content": "content_problem_links", "book": "book_problem_links", "note": "note_problem_links"}[entity]
    key = {"content": "content_id", "book": "book_id", "note": "reading_note_id"}[entity]
    rows = query_all(
        f"SELECT p.name FROM problems p JOIN {table} l ON p.id = l.problem_id WHERE l.{key} = ? AND p.deleted_at IS NULL",
        (entity_id,),
    )
    return [r["name"] for r in rows]


@app.route("/")
def index():
    stats = {
        "need_reason": query_one("SELECT COUNT(*) c FROM contents WHERE deleted_at IS NULL AND status='待补充'")["c"],
        "need_judge": query_one("SELECT COUNT(*) c FROM contents WHERE deleted_at IS NULL AND status='待判断'")["c"],
        "ready": query_one("SELECT COUNT(*) c FROM experiments WHERE deleted_at IS NULL AND status='待实践'")["c"],
        "doing": query_one("SELECT COUNT(*) c FROM experiments WHERE deleted_at IS NULL AND status='实践中'")["c"],
        "review": query_one("SELECT COUNT(*) c FROM experiments WHERE deleted_at IS NULL AND status='等待复盘'")["c"],
    }
    recent = query_all(
        "SELECT id,title,reusable_conclusion,updated_at FROM experiments WHERE deleted_at IS NULL AND reusable_conclusion<>'' ORDER BY updated_at DESC LIMIT 5"
    )
    reading = query_all("SELECT id,title,author FROM books WHERE deleted_at IS NULL AND reading_status='阅读中' ORDER BY updated_at DESC LIMIT 5")
    return render_template("index.html", stats=stats, recent=recent, reading=reading)


@app.route("/capture", methods=["GET", "POST"])
def capture():
    if request.method == "POST":
        url = form_value("url")
        wants_practice = form_value("wants_practice") == "1"
        extracted = extract_url(url) if form_value("try_extract") == "1" else {"status": "仅保存链接", "error": "", "title": "", "raw_text": "", "summary": ""}
        title = form_value("title") or extracted["title"] or urlparse(url).netloc or "未命名内容"
        save_reason = form_value("save_reason")
        status = content_status(save_reason, wants_practice)
        content_id = execute(
            """
            INSERT INTO contents(content_type,title,url,author,source_platform,published_at,raw_text,summary,extraction_status,
              extraction_error,content_source,user_reflection,save_reason,problem_statement,intended_use,status,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                form_value("content_type") or "article",
                title,
                url,
                form_value("author"),
                form_value("source_platform"),
                form_value("published_at"),
                form_value("raw_text") or extracted["raw_text"],
                extracted["summary"],
                extracted["status"],
                extracted["error"],
                "url" if url else "manual",
                form_value("user_reflection"),
                save_reason,
                form_value("problem_statement"),
                form_value("intended_use"),
                status,
                now(),
                now(),
            ),
        )
        problem_id = form_value("problem_id")
        if problem_id:
            execute("INSERT OR IGNORE INTO content_problem_links(content_id, problem_id, created_at) VALUES(?,?,?)", (content_id, problem_id, now()))
        save_attachments("content", content_id)
        if wants_practice:
            execute(
                "INSERT INTO experiments(title,content_id,problem_id,goal,current_problem,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)",
                (title, content_id, problem_id or None, form_value("intended_use"), form_value("problem_statement"), "待实践", now(), now()),
            )
        flash(f"已保存内容。提取状态：{extracted['status']}", "success" if extracted["status"] != "读取失败" else "warning")
        return redirect(url_for("content_detail", content_id=content_id))
    problems = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY updated_at DESC")
    return render_template("capture.html", problems=problems, statuses=CONTENT_STATUSES)


@app.route("/contents")
def contents():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    params = []
    where = ["deleted_at IS NULL"]
    if q:
        where.append("(title LIKE ? OR raw_text LIKE ? OR summary LIKE ? OR user_reflection LIKE ? OR save_reason LIKE ? OR problem_statement LIKE ?)")
        params.extend([f"%{q}%"] * 6)
    if status:
        where.append("status=?")
        params.append(status)
    rows = query_all(f"SELECT * FROM contents WHERE {' AND '.join(where)} ORDER BY updated_at DESC", tuple(params))
    return render_template("contents.html", rows=rows, q=q, status=status, statuses=CONTENT_STATUSES)


@app.route("/contents/<int:content_id>", methods=["GET", "POST"])
def content_detail(content_id: int):
    if request.method == "POST":
        execute(
            """
            UPDATE contents SET content_type=?,title=?,url=?,author=?,source_platform=?,published_at=?,raw_text=?,summary=?,tools=?,methods=?,
              suggested_category=?,applicable_scenarios=?,user_reflection=?,save_reason=?,problem_statement=?,intended_use=?,status=?,updated_at=?
            WHERE id=?
            """,
            (
                form_value("content_type"),
                form_value("title"),
                form_value("url"),
                form_value("author"),
                form_value("source_platform"),
                form_value("published_at"),
                form_value("raw_text"),
                form_value("summary"),
                form_value("tools"),
                form_value("methods"),
                form_value("suggested_category"),
                form_value("applicable_scenarios"),
                form_value("user_reflection"),
                form_value("save_reason"),
                form_value("problem_statement"),
                form_value("intended_use"),
                form_value("status"),
                now(),
                content_id,
            ),
        )
        problem_id = form_value("problem_id")
        if problem_id:
            execute("INSERT OR IGNORE INTO content_problem_links(content_id, problem_id, created_at) VALUES(?,?,?)", (content_id, problem_id, now()))
        save_attachments("content", content_id)
        flash("内容已更新。", "success")
        return redirect(url_for("content_detail", content_id=content_id))
    row = query_one("SELECT * FROM contents WHERE id=? AND deleted_at IS NULL", (content_id,))
    problems = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    linked = existing_problem_links("content", content_id)
    attachments = query_all("SELECT * FROM attachments WHERE entity_type='content' AND entity_id=? ORDER BY created_at DESC", (content_id,))
    experiments = query_all("SELECT * FROM experiments WHERE content_id=? AND deleted_at IS NULL ORDER BY updated_at DESC", (content_id,))
    return render_template("content_detail.html", row=row, problems=problems, linked=linked, attachments=attachments, experiments=experiments, statuses=CONTENT_STATUSES)


@app.post("/contents/<int:content_id>/experiment")
def content_to_experiment(content_id: int):
    row = query_one("SELECT * FROM contents WHERE id=? AND deleted_at IS NULL", (content_id,))
    if not row:
        flash("内容不存在。", "error")
        return redirect(url_for("contents"))
    problem = query_one("SELECT problem_id FROM content_problem_links WHERE content_id=? LIMIT 1", (content_id,))
    exp_id = execute(
        "INSERT INTO experiments(title,content_id,problem_id,goal,current_problem,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)",
        (row["title"], content_id, problem["problem_id"] if problem else None, row["intended_use"], row["problem_statement"], "待实践", now(), now()),
    )
    execute("UPDATE contents SET status='待实践', updated_at=? WHERE id=?", (now(), content_id))
    flash("已转为实践，请补充最小测试和成功标准。", "success")
    return redirect(url_for("experiment_detail", experiment_id=exp_id))


@app.route("/problems", methods=["GET", "POST"])
def problems():
    if request.method == "POST":
        execute(
            "INSERT INTO problems(name,description,problem_type,importance,status,current_conclusion,next_action,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (form_value("name"), form_value("description"), form_value("problem_type"), form_value("importance"), form_value("status") or "待处理", form_value("current_conclusion"), form_value("next_action"), now(), now()),
        )
        flash("问题已创建。", "success")
        return redirect(url_for("problems"))
    rows = query_all("SELECT * FROM problems WHERE deleted_at IS NULL ORDER BY updated_at DESC")
    return render_template("problems.html", rows=rows, statuses=PROBLEM_STATUSES)


@app.route("/problems/<int:problem_id>", methods=["GET", "POST"])
def problem_detail(problem_id: int):
    if request.method == "POST":
        execute(
            "UPDATE problems SET name=?,description=?,problem_type=?,importance=?,status=?,current_conclusion=?,next_action=?,updated_at=? WHERE id=?",
            (form_value("name"), form_value("description"), form_value("problem_type"), form_value("importance"), form_value("status"), form_value("current_conclusion"), form_value("next_action"), now(), problem_id),
        )
        flash("问题已更新。", "success")
        return redirect(url_for("problem_detail", problem_id=problem_id))
    row = query_one("SELECT * FROM problems WHERE id=? AND deleted_at IS NULL", (problem_id,))
    contents = query_all("SELECT c.* FROM contents c JOIN content_problem_links l ON c.id=l.content_id WHERE l.problem_id=? AND c.deleted_at IS NULL", (problem_id,))
    books = query_all("SELECT b.* FROM books b JOIN book_problem_links l ON b.id=l.book_id WHERE l.problem_id=? AND b.deleted_at IS NULL", (problem_id,))
    notes = query_all("SELECT n.*, b.title book_title FROM reading_notes n LEFT JOIN books b ON b.id=n.book_id JOIN note_problem_links l ON n.id=l.reading_note_id WHERE l.problem_id=? AND n.deleted_at IS NULL", (problem_id,))
    experiments = query_all("SELECT * FROM experiments WHERE problem_id=? AND deleted_at IS NULL ORDER BY updated_at DESC", (problem_id,))
    return render_template("problem_detail.html", row=row, contents=contents, books=books, notes=notes, experiments=experiments, statuses=PROBLEM_STATUSES)


@app.route("/experiments", methods=["GET", "POST"])
def experiments():
    if request.method == "POST":
        exp_id = save_experiment()
        if not exp_id:
            return redirect(url_for("experiments"))
        flash("实践已创建。", "success")
        return redirect(url_for("experiment_detail", experiment_id=exp_id))
    rows = query_all(
        "SELECT e.*, c.title content_title, b.title book_title, p.name problem_name FROM experiments e LEFT JOIN contents c ON c.id=e.content_id LEFT JOIN books b ON b.id=e.book_id LEFT JOIN problems p ON p.id=e.problem_id WHERE e.deleted_at IS NULL ORDER BY e.updated_at DESC"
    )
    contents_rows = query_all("SELECT id,title FROM contents WHERE deleted_at IS NULL ORDER BY title")
    books_rows = query_all("SELECT id,title FROM books WHERE deleted_at IS NULL ORDER BY title")
    notes_rows = query_all("SELECT id,source_idea FROM reading_notes WHERE deleted_at IS NULL ORDER BY updated_at DESC")
    problems_rows = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    return render_template("experiments.html", rows=rows, statuses=EXPERIMENT_STATUSES, contents_rows=contents_rows, books_rows=books_rows, notes_rows=notes_rows, problems_rows=problems_rows)


def validate_experiment_status(status: str, actual_result: str, final_decision: str, reusable_conclusion: str, issues_found: str) -> str | None:
    if status == "已验证" and (not actual_result or not reusable_conclusion):
        return "状态改为“已验证”前，请填写实际结果和可复用结论。"
    if status == "不适用" and (not issues_found and not final_decision):
        return "状态改为“不适用”前，请填写不适用原因，可放在“遇到的问题”或“最终决定”。"
    if status in ["已验证", "不适用"] and not final_decision:
        return "结束实践前，请填写最终决定。"
    return None


def save_experiment(experiment_id: int | None = None) -> int:
    status = form_value("status") or "待实践"
    error = validate_experiment_status(status, form_value("actual_result"), form_value("final_decision"), form_value("reusable_conclusion"), form_value("issues_found"))
    if error:
        flash(error, "error")
        return experiment_id or 0
    values = (
        form_value("title"),
        form_value("content_id") or None,
        form_value("book_id") or None,
        form_value("reading_note_id") or None,
        form_value("problem_id") or None,
        form_value("goal"),
        form_value("current_problem"),
        form_value("minimum_test"),
        form_value("success_criteria"),
        form_value("estimated_effort"),
        form_value("started_at"),
        form_value("actual_action"),
        form_value("actual_result"),
        form_value("issues_found"),
        form_value("is_effective"),
        form_value("final_decision"),
        form_value("reusable_conclusion"),
        form_value("next_action"),
        status,
        now(),
    )
    if experiment_id:
        execute(
            """
            UPDATE experiments SET title=?,content_id=?,book_id=?,reading_note_id=?,problem_id=?,goal=?,current_problem=?,minimum_test=?,
              success_criteria=?,estimated_effort=?,started_at=?,actual_action=?,actual_result=?,issues_found=?,is_effective=?,
              final_decision=?,reusable_conclusion=?,next_action=?,status=?,updated_at=? WHERE id=?
            """,
            values + (experiment_id,),
        )
        return experiment_id
    return execute(
        """
        INSERT INTO experiments(title,content_id,book_id,reading_note_id,problem_id,goal,current_problem,minimum_test,success_criteria,
          estimated_effort,started_at,actual_action,actual_result,issues_found,is_effective,final_decision,reusable_conclusion,next_action,status,created_at,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        values[:-1] + (now(), now()),
    )


@app.route("/experiments/<int:experiment_id>", methods=["GET", "POST"])
def experiment_detail(experiment_id: int):
    if request.method == "POST":
        saved_id = save_experiment(experiment_id)
        if saved_id:
            flash("实践已更新。", "success")
        return redirect(url_for("experiment_detail", experiment_id=experiment_id))
    row = query_one("SELECT * FROM experiments WHERE id=? AND deleted_at IS NULL", (experiment_id,))
    if not row:
        flash("实践记录不存在或已删除。", "error")
        return redirect(url_for("experiments"))
    contents_rows = query_all("SELECT id,title FROM contents WHERE deleted_at IS NULL ORDER BY title")
    books_rows = query_all("SELECT id,title FROM books WHERE deleted_at IS NULL ORDER BY title")
    notes_rows = query_all("SELECT id,source_idea FROM reading_notes WHERE deleted_at IS NULL ORDER BY updated_at DESC")
    problems_rows = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    return render_template("experiment_detail.html", row=row, statuses=EXPERIMENT_STATUSES, contents_rows=contents_rows, books_rows=books_rows, notes_rows=notes_rows, problems_rows=problems_rows)


@app.route("/books", methods=["GET", "POST"])
def books():
    if request.method == "POST":
        book_id = execute(
            """
            INSERT INTO books(title,author,isbn,publisher,published_at,description,category,problem_statement,relationship_status,reading_status,
              priority,is_read,start_date,finish_date,user_note,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (form_value("title"), form_value("author"), form_value("isbn"), form_value("publisher"), form_value("published_at"), form_value("description"), form_value("category"), form_value("problem_statement"), form_value("relationship_status"), form_value("reading_status") or "未开始", form_value("priority"), 1 if form_value("is_read") else 0, form_value("start_date"), form_value("finish_date"), form_value("user_note"), now(), now()),
        )
        problem_id = form_value("problem_id")
        if problem_id:
            execute("INSERT OR IGNORE INTO book_problem_links(book_id, problem_id, created_at) VALUES(?,?,?)", (book_id, problem_id, now()))
        flash("书籍已保存。", "success")
        return redirect(url_for("books"))
    rows = query_all("SELECT * FROM books WHERE deleted_at IS NULL ORDER BY updated_at DESC")
    problems_rows = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    return render_template("books.html", rows=rows, problems=problems_rows, relationships=BOOK_RELATIONSHIPS, reading_statuses=READING_STATUSES, priorities=PRIORITIES)


@app.route("/books/<int:book_id>", methods=["GET", "POST"])
def book_detail(book_id: int):
    if request.method == "POST":
        execute(
            """
            UPDATE books SET title=?,author=?,isbn=?,publisher=?,published_at=?,description=?,category=?,problem_statement=?,
              relationship_status=?,reading_status=?,priority=?,is_read=?,start_date=?,finish_date=?,user_note=?,updated_at=? WHERE id=?
            """,
            (form_value("title"), form_value("author"), form_value("isbn"), form_value("publisher"), form_value("published_at"), form_value("description"), form_value("category"), form_value("problem_statement"), form_value("relationship_status"), form_value("reading_status"), form_value("priority"), 1 if form_value("is_read") else 0, form_value("start_date"), form_value("finish_date"), form_value("user_note"), now(), book_id),
        )
        problem_id = form_value("problem_id")
        if problem_id:
            execute("INSERT OR IGNORE INTO book_problem_links(book_id, problem_id, created_at) VALUES(?,?,?)", (book_id, problem_id, now()))
        flash("书籍已更新。", "success")
        return redirect(url_for("book_detail", book_id=book_id))
    row = query_one("SELECT * FROM books WHERE id=? AND deleted_at IS NULL", (book_id,))
    problems_rows = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    linked = existing_problem_links("book", book_id)
    notes = query_all("SELECT * FROM reading_notes WHERE book_id=? AND deleted_at IS NULL ORDER BY updated_at DESC", (book_id,))
    return render_template("book_detail.html", row=row, problems=problems_rows, linked=linked, notes=notes, relationships=BOOK_RELATIONSHIPS, reading_statuses=READING_STATUSES, priorities=PRIORITIES)


@app.route("/books/import", methods=["POST"])
def import_books():
    file = request.files.get("book_file")
    if not file or not file.filename:
        flash("请选择 CSV 或 Excel 文件。", "error")
        return redirect(url_for("books"))
    rows = read_book_file(file)
    imported = skipped = 0
    for item in rows:
        title = (item.get("书名") or item.get("title") or "").strip()
        if not title:
            skipped += 1
            continue
        author = (item.get("作者") or item.get("author") or "").strip()
        duplicate = query_one("SELECT id FROM books WHERE deleted_at IS NULL AND title=? AND COALESCE(author,'')=?", (title, author))
        if duplicate:
            skipped += 1
            continue
        execute(
            "INSERT INTO books(title,author,isbn,is_read,relationship_status,user_note,reading_status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (title, author, item.get("ISBN") or item.get("isbn"), 1 if str(item.get("是否读过", "")).strip() in ["是", "1", "true", "True"] else 0, item.get("当前兴趣") or "", item.get("备注") or "", "未开始", now(), now()),
        )
        imported += 1
    flash(f"导入完成：新增 {imported} 本，跳过 {skipped} 行或重复书籍。", "success")
    return redirect(url_for("books"))


def read_book_file(file) -> list[dict]:
    filename = file.filename.lower()
    if filename.endswith(".csv"):
        text = file.stream.read().decode("utf-8-sig")
        return list(csv.DictReader(text.splitlines()))
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v or "").strip() for v in rows[0]]
        return [{headers[i]: (v or "") for i, v in enumerate(row)} for row in rows[1:] if any(row)]
    return []


@app.route("/notes", methods=["GET", "POST"])
def notes():
    if request.method == "POST":
        note_id = execute(
            "INSERT INTO reading_notes(book_id,chapter_or_page,source_idea,user_understanding,related_experience,planned_action,convert_to_experiment,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (form_value("book_id") or None, form_value("chapter_or_page"), form_value("source_idea"), form_value("user_understanding"), form_value("related_experience"), form_value("planned_action"), 1 if form_value("convert_to_experiment") else 0, now(), now()),
        )
        problem_id = form_value("problem_id")
        if problem_id:
            execute("INSERT OR IGNORE INTO note_problem_links(reading_note_id, problem_id, created_at) VALUES(?,?,?)", (note_id, problem_id, now()))
        if form_value("convert_to_experiment"):
            execute(
                "INSERT INTO experiments(title,book_id,reading_note_id,problem_id,goal,minimum_test,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
                (form_value("source_idea")[:80] or "读书感悟实践", form_value("book_id") or None, note_id, problem_id or None, form_value("planned_action"), form_value("planned_action"), "待实践", now(), now()),
            )
        flash("读书感悟已保存。", "success")
        return redirect(url_for("notes"))
    rows = query_all("SELECT n.*, b.title book_title FROM reading_notes n LEFT JOIN books b ON b.id=n.book_id WHERE n.deleted_at IS NULL ORDER BY n.updated_at DESC")
    books_rows = query_all("SELECT id,title FROM books WHERE deleted_at IS NULL ORDER BY title")
    problems_rows = query_all("SELECT id,name FROM problems WHERE deleted_at IS NULL ORDER BY name")
    return render_template("notes.html", rows=rows, books=books_rows, problems=problems_rows)


@app.route("/export")
def export_page():
    return render_template("export.html")


@app.post("/export/markdown/<entity>/<int:entity_id>")
def export_markdown(entity: str, entity_id: int):
    if entity == "content":
        row = query_one("SELECT * FROM contents WHERE id=?", (entity_id,))
        links = "\n".join(f"- [[{p}]]" for p in existing_problem_links("content", entity_id))
        body = f"""# {row['title']}

## 来源

链接：{row['url'] or ''}
作者：{row['author'] or ''}
平台：{row['source_platform'] or ''}
发布时间：{row['published_at'] or ''}

## AI 摘要

> 此部分为 AI 辅助信息

{row['summary'] or ''}

## 我为什么收藏

{row['save_reason'] or ''}

## 哪一点打动了我

{row['user_reflection'] or ''}

## 它可能解决的问题

{row['problem_statement'] or ''}

## 我准备怎么使用

{row['intended_use'] or ''}

## 当前状态

{row['status'] or ''}

## 关联问题

{links}
"""
        path = write_markdown("内容收藏", row["title"], body)
    elif entity == "experiment":
        row = query_one("SELECT e.*, p.name problem_name FROM experiments e LEFT JOIN problems p ON p.id=e.problem_id WHERE e.id=?", (entity_id,))
        body = f"""# {row['title']}

## 要解决的问题

{f"[[{row['problem_name']}]]" if row['problem_name'] else row['current_problem'] or ''}

## 验证目标

{row['goal'] or ''}

## 最小测试

{row['minimum_test'] or ''}

## 实际操作

{row['actual_action'] or ''}

## 实际结果

{row['actual_result'] or ''}

## 最终决定

{row['final_decision'] or ''}

## 可复用结论

{row['reusable_conclusion'] or ''}

## 适用场景

{row['issues_found'] or ''}

## 下一步行动

{row['next_action'] or ''}
"""
        path = write_markdown("实践记录", row["title"], body)
    elif entity == "note":
        row = query_one("SELECT n.*, b.title book_title FROM reading_notes n LEFT JOIN books b ON b.id=n.book_id WHERE n.id=?", (entity_id,))
        links = "\n".join(f"- [[{p}]]" for p in existing_problem_links("note", entity_id))
        body = f"""# {row['book_title'] or '未关联书籍'}：读书感悟

## 章节或页码

{row['chapter_or_page'] or ''}

## 原文观点

{row['source_idea'] or ''}

## 我的理解

{row['user_understanding'] or ''}

## 关联经历

{row['related_experience'] or ''}

## 我准备怎么用

{row['planned_action'] or ''}

## 关联问题

{links}
"""
        path = write_markdown("读书感悟", row["book_title"] or "读书感悟", body)
    else:
        flash("不支持的导出类型。", "error")
        return redirect(request.referrer or url_for("index"))
    flash(f"已导出 Markdown：{path}", "success")
    return redirect(request.referrer or url_for("export_page"))


@app.route("/export/json")
def export_json():
    data = {}
    for table in ["contents", "books", "reading_notes", "problems", "experiments", "content_problem_links", "book_problem_links", "note_problem_links", "attachments"]:
        data[table] = [dict(r) for r in query_all(f"SELECT * FROM {table}")]
    path = EXPORTS_DIR / "json" / f"knowledge_export_{stamp()}.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    flash(f"已导出 JSON：{path}", "success")
    return send_file(path, as_attachment=True)


@app.route("/export/csv/<table>")
def export_csv(table: str):
    if table not in ["books", "contents"]:
        flash("只支持导出 books 或 contents。", "error")
        return redirect(url_for("export_page"))
    rows = query_all(f"SELECT * FROM {table} WHERE deleted_at IS NULL")
    path = EXPORTS_DIR / "csv" / f"{table}_{stamp()}.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        if rows:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows([dict(r) for r in rows])
    flash(f"已导出 CSV：{path}", "success")
    return send_file(path, as_attachment=True)


@app.post("/backup")
def backup():
    init_db()
    path = BACKUPS_DIR / f"knowledge_backup_{stamp()}.db"
    shutil.copy2(DB_PATH, path)
    flash(f"已备份数据库：{path}", "success")
    return redirect(url_for("export_page"))


@app.post("/restore")
def restore():
    file = request.files.get("backup_file")
    if not file or not file.filename:
        flash("请选择要恢复的 .db 文件。", "error")
        return redirect(url_for("export_page"))
    backup_path = BACKUPS_DIR / f"before_restore_{stamp()}.db"
    if DB_PATH.exists():
        shutil.copy2(DB_PATH, backup_path)
    file.save(DB_PATH)
    flash(f"已恢复数据库。恢复前备份：{backup_path}", "warning")
    return redirect(url_for("index"))


@app.post("/delete/<entity>/<int:entity_id>")
def soft_delete(entity: str, entity_id: int):
    table = {"content": "contents", "book": "books", "note": "reading_notes", "problem": "problems", "experiment": "experiments"}.get(entity)
    if not table:
        flash("不支持的删除类型。", "error")
        return redirect(url_for("index"))
    execute(f"UPDATE {table} SET deleted_at=?, updated_at=? WHERE id=?", (now(), now(), entity_id))
    flash("已移入回收状态，未永久删除。", "warning")
    return redirect(url_for("index"))


@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        for key in ["ai_provider", "ai_api_url", "ai_api_key", "ai_model"]:
            execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, form_value(key)))
        flash("设置已保存。AI 仍只作为建议，不会自动覆盖人工判断。", "success")
        return redirect(url_for("settings"))
    rows = {r["key"]: r["value"] for r in query_all("SELECT key,value FROM settings")}
    return render_template("settings.html", settings=rows)


@app.route("/api/search")
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    like = f"%{q}%"
    results = []
    for r in query_all("SELECT id,title,'内容' kind FROM contents WHERE deleted_at IS NULL AND (title LIKE ? OR raw_text LIKE ? OR summary LIKE ? OR user_reflection LIKE ? OR save_reason LIKE ?) LIMIT 8", (like, like, like, like, like)):
        results.append({"kind": r["kind"], "title": r["title"], "url": url_for("content_detail", content_id=r["id"])})
    for r in query_all("SELECT id,name title,'问题' kind FROM problems WHERE deleted_at IS NULL AND (name LIKE ? OR description LIKE ? OR current_conclusion LIKE ?) LIMIT 8", (like, like, like)):
        results.append({"kind": r["kind"], "title": r["title"], "url": url_for("problem_detail", problem_id=r["id"])})
    for r in query_all("SELECT id,title,'书籍' kind FROM books WHERE deleted_at IS NULL AND (title LIKE ? OR author LIKE ? OR problem_statement LIKE ?) LIMIT 8", (like, like, like)):
        results.append({"kind": r["kind"], "title": r["title"], "url": url_for("book_detail", book_id=r["id"])})
    return jsonify(results)


@app.context_processor
def inject_globals():
    return {"content_statuses": CONTENT_STATUSES, "experiment_statuses": EXPERIMENT_STATUSES}


if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=5000, debug=False)
